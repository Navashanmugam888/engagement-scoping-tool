"""
FTE Efforts Calculator

Calculates role-based FTE (Full-Time Equivalent) effort allocation.
Uses SUMPRODUCT formula: hours × role allocation percentage for each tier.

Formula pattern from Excel:
=SUMPRODUCT($I$6:$I$22, J6:J22)

Where:
- $I$6:$I$22 = Hours from Effort Estimation for each tier
- J6:J22 = Role allocation percentage (0-1) for each tier
- Result = Total FTE hours for that role across all tiers
"""

from pathlib import Path
import sys
import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

from backend.config import EXCEL_FILE


class FTEEffortsCalculator:
    """Calculate role-based FTE effort allocation"""
    
    def __init__(self):
        """Load tier definitions and role allocation data from App Tiers Definition sheet"""
        self.tiers_data = {}  # Dict with row_index -> {hours, roles: {role -> allocation}}
        self.roles = []
        self.tier_hours = {}
        self.category_tier_mapping = {}  # Maps category name to tier row index
        
        self._load_app_tiers_data()
    
    def _load_app_tiers_data(self):
        """Load tier and role allocation data from App Tiers Definition sheet"""
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb['App Tiers Definition']
        
        # Structure:
        # Row 4 has location (USA/India), Row 5 has role name
        # This creates full role names like "PM USA", "PM India", "Architect USA", etc.
        # Column I = Hours (rows 6-22)
        # Columns J onwards = Role data (rows 6-22)
        # Row 24 contains SUMPRODUCT formulas (totals)
        
        # Load role names from rows 4-5, starting from column J (column 10)
        col_index = 10  # J is column 10
        while True:
            location_cell = ws.cell(row=4, column=col_index)
            role_cell = ws.cell(row=5, column=col_index)
            
            if not role_cell.value:
                break
            
            # Create full role name: "Role Location"
            role_name = str(role_cell.value).strip()
            location = str(location_cell.value).strip() if location_cell.value else ""
            
            if location:
                full_role_name = f"{role_name} {location}"
            else:
                full_role_name = role_name
            
            self.roles.append(full_role_name)
            col_index += 1
        
        # Load tier/category data for all 17 rows (rows 6-22)
        # These include both tier definitions (6-10) and category rows (11-22)
        for row_idx, row in enumerate(range(6, 23)):  # Rows 6-22 (17 rows)
            hours = ws.cell(row=row, column=9).value  # Column I = Hours
            
            if hours is not None:
                # Initialize tier data
                if row_idx not in self.tiers_data:
                    self.tiers_data[row_idx] = {
                        'hours': hours,
                        'row': row,
                        'roles': {}
                    }
                
                # Load role allocation percentages for this tier/category
                col_index = 10  # J is column 10
                for role in self.roles:
                    allocation = ws.cell(row=row, column=col_index).value
                    if allocation is not None:
                        # Convert to float if it's a percentage string
                        if isinstance(allocation, str) and '%' in allocation:
                            allocation = float(allocation.strip('%')) / 100
                        self.tiers_data[row_idx]['roles'][role] = float(allocation) if allocation else 0
                    col_index += 1
        
        print(f"Loaded {len(self.roles)} roles: {self.roles}")
        print(f"Loaded {len(self.tiers_data)} tier/category rows (6-22)")
    
    def _load_category_mapping(self):
        """Load the mapping from categories to tier rows using the Hours formulas"""
        wb = load_workbook(EXCEL_FILE, data_only=False)
        ws = wb['App Tiers Definition']
        
        # Extract effort estimation references from Hours column formulas
        # I6 = 'Effort Estimation'!G5, I7 = 'Effort Estimation'!G14, etc.
        # Parse these to identify which category each row represents
        
        for row in range(6, 23):
            formula = ws[f'I{row}'].value
            if formula and formula.startswith("='Effort Estimation'!"):
                # Extract cell reference like G5
                cell_ref = formula.split('!')[-1].strip("'")
                # Now we need to find which category has this cell
                # For now, use row - 6 as tier index (0-16)
                tier_idx = row - 6
                self.category_tier_mapping[tier_idx] = cell_ref
    
    def calculate_role_fte_from_effort(self, effort_estimation: dict, selected_roles: list = None) -> dict:
        """
        Calculate FTE hours for each role using effort estimation output
        
        Uses SUMPRODUCT: Sum of (Category_Hours × Role_Allocation_%)
        For each category, multiply the effort hours by the role's allocation percentage,
        then sum across all 17 rows (rows 6-22 in App Tiers Definition).
        
        Excel Formula Pattern: =SUMPRODUCT($I6:$I22, J6:J22)
        Where:
        - I6:I22 = Hours (17 rows, includes tier template + categories)
        - J6:J22 = Role allocation percentage for that row
        
        Args:
            effort_estimation: dict from EffortCalculator.calculate_effort()
                              with category_name -> {'final_estimate': hours, ...} mapping
            selected_roles: list of selected role names (None = all roles)
        
        Returns:
            dict with role_name -> fte_hours mapping
        """
        role_fte = {}
        roles_to_calculate = selected_roles if selected_roles else self.roles
        
        # For each role, calculate SUMPRODUCT across all 17 rows (rows 6-22)
        for role in roles_to_calculate:
            if role not in self.roles:
                continue
            
            total_fte = 0.0
            
            # Iterate through each row 0-16 (representing rows 6-22 in Excel)
            for row_idx in range(17):  # 17 rows total (rows 6-22)
                if row_idx not in self.tiers_data:
                    continue
                
                tier_data = self.tiers_data[row_idx]
                
                # Hours for this row from App Tiers Definition (column I)
                # This value is read from the App Tiers sheet and may reference Effort Estimation
                hours_value = tier_data['hours']
                
                # Get role allocation percentage for this row
                role_allocation = tier_data['roles'].get(role, 0.0)
                
                # SUMPRODUCT: hours × allocation, then sum across all rows
                total_fte += hours_value * role_allocation
            
            role_fte[role] = total_fte
        
        return role_fte
    
    def get_role_allocation_matrix(self) -> pd.DataFrame:
        """
        Get the complete role allocation matrix as DataFrame
        
        Returns:
            DataFrame with tiers as rows and roles as columns
        """
        data = []
        for tier_name in sorted(self.tiers_data.keys()):
            tier_data = self.tiers_data[tier_name]
            row = {'Tier': tier_name, 'Hours': tier_data['hours']}
            
            for role in self.roles:
                row[role] = tier_data['roles'].get(role, 0)
            
            data.append(row)
        
        return pd.DataFrame(data)
    
    def generate_fte_summary(self, effort_estimation: dict, selected_roles: list = None) -> dict:
        """
        Generate complete FTE summary for all selected roles
        
        Args:
            effort_estimation: Output from EffortCalculator.calculate_effort()
            selected_roles: List of selected role names
        
        Returns:
            dict with:
            - role_fte_hours: dict of role -> fte_hours
            - role_fte_days: dict of role -> fte_days (hours/8)
            - role_fte_summary: list of dicts with role details
        """
        # Map effort estimation by tier
        tier_efforts = {}
        for category, cat_data in effort_estimation.items():
            # Get tier name from effort_estimation summary
            # For now, assume we need to get this from scope_result
            pass
        
        # Calculate role FTE
        role_fte_hours = self.calculate_role_fte(tier_efforts, selected_roles)
        
        # Convert to days
        role_fte_days = {role: hours / 8 for role, hours in role_fte_hours.items()}
        
        # Create summary
        summary = []
        for role in (selected_roles if selected_roles else self.roles):
            if role in role_fte_hours:
                summary.append({
                    'role': role,
                    'fte_hours': role_fte_hours[role],
                    'fte_days': role_fte_days[role],
                    'fte_months': role_fte_days[role] / 30
                })
        
        return {
            'role_fte_hours': role_fte_hours,
            'role_fte_days': role_fte_days,
            'role_fte_summary': summary
        }

