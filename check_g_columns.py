from openpyxl import load_workbook

# Load Excel with formulas
wb = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=False)
ws = wb['Effort Estimation']

# Check row 84 G column formula
print('Row 84 - Historical Data category:')
print(f'  G84 formula: {ws["G84"].value}')

# Check what G84 should sum
print('\nRows 84-88, Column G:')
for row in range(84, 89):
    cell = ws[f'G{row}']
    print(f'  G{row}: {cell.value}')

# With data_only
wb2 = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=True)
ws2 = wb2['Effort Estimation']

print('\nWith calculated values:')
print(f'  G84: {ws2["G84"].value}')
print(f'  G85: {ws2["G85"].value}')
print(f'  G86: {ws2["G86"].value}')
print(f'  G87: {ws2["G87"].value}')
print(f'  G88: {ws2["G88"].value}')
