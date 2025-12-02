from openpyxl import load_workbook

wb = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=False)
ws = wb['Scope Definition']

print('Scope Definition sheet:')
print(f'D13: {ws["D13"].value}')
print(f'D54: {ws["D54"].value}')

# Find the row with these
print('\nRow 13:')
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    print(f'  {col}13: {ws[f"{col}13"].value}')

print('\nRow 54:')
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    print(f'  {col}54: {ws[f"{col}54"].value}')

# Check what the formula references
wb2 = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=True)
ws2 = wb2['Scope Definition']

print('\n\nWith calculated values:')
print(f'D13: {ws2["D13"].value}')
print(f'D54: {ws2["D54"].value}')
