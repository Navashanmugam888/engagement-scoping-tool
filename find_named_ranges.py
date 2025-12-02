from openpyxl import load_workbook

wb = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=False)
ws = wb['Effort Estimation']

# Search for EngagementWeightage definition
print('Searching for EngagementWeightage name...')
for name in wb.named_ranges:
    print(f'Named range: {name}')

# Try to find it in formulas
print('\n\nSearching for cells containing "EngagementWeightage"...')
for row in range(1, 100):
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws[f'{col}{row}']
        if cell.value and isinstance(cell.value, str) and 'EngagementWeightage' in cell.value:
            print(f'{col}{row}: {cell.value}')

# Also check HistData
print('\n\nSearching for cells containing "HistData"...')
for row in range(1, 100):
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws[f'{col}{row}']
        if cell.value and isinstance(cell.value, str) and 'HistData' in cell.value:
            print(f'{col}{row}: {cell.value}')
