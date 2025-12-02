"""Check Excel for grey cells and metric names"""
from openpyxl import load_workbook
from pathlib import Path

wb = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=False)
ws = wb['Scope Definition']

print("Row | Metric Name | InScope | Details | Grey?")
print("="*80)

for row in range(6, 103):
    name = ws.cell(row, 2).value
    inscope = ws.cell(row, 3).value
    details_val = ws.cell(row, 4).value
    
    # Check if details cell is grey
    details_cell = ws.cell(row, 4)
    is_grey = False
    if details_cell.fill and details_cell.fill.start_color:
        color = details_cell.fill.start_color.index
        is_grey = color != '00000000' and color != 'FFFFFFFF' and color != 'FF000000'
    
    if name and isinstance(name, str):
        name_str = str(name).strip()
        grey_marker = " [GREY]" if is_grey else ""
        print(f"{row:3d} | {name_str:45s} | {inscope:7s} | {details_val if details_val else '':10s} | {grey_marker}")
