import openpyxl

wb = openpyxl.load_workbook('Engagement Scoping Tool - FCC.xlsx')
print('Sheet names:', wb.sheetnames)

ws = wb['Scope Definition']
print(f'\nMax row: {ws.max_row}')
print(f'Max column: {ws.max_column}')

print('\n=== ALL DATA IN SCOPE DEFINITION ===\n')
for row_idx in range(1, ws.max_row+1):
    col_a = ws.cell(row=row_idx, column=1).value
    col_b = ws.cell(row=row_idx, column=2).value
    col_c = ws.cell(row=row_idx, column=3).value
    
    if col_a or col_b or col_c:
        print(f'Row {row_idx}: "{col_a}" | "{col_b}" | "{col_c}"')
