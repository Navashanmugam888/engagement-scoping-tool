import openpyxl

excel_file = 'C:/Users/NavashanmugamAsokan/Desktop/Projects/Engagement Scoping Tool/documents/Engagement Scoping Tool - FCC.xlsx'

wb = openpyxl.load_workbook(excel_file, data_only=False)
ws = wb['Effort Estimation']

print('=== CATEGORY HEADER TIER ADJUSTMENT FORMULAS ===')
print()

category_rows = [
    (5, 'Project Initiation and Planning'),
    (14, 'Creating and Managing EPM Cloud Infrastructure'),
    (18, 'Requirement Gathering, Read back and Client Sign-off'),
    (28, 'Design'),
    (35, 'Build and Configure FCC'),
    (53, 'Setup Application Features'),
    (70, 'Application Customization'),
    (74, 'Calculations'),
    (80, 'Security'),
    (84, 'Historical Data'),
    (90, 'Integrations'),
    (97, 'Reporting'),
    (106, 'Automations'),
    (112, 'Testing/Training'),
    (119, 'Transition'),
    (123, 'Documentations'),
    (128, 'Change Management'),
]

for row_idx, category_name in category_rows:
    cell_f = ws.cell(row_idx, 6)  # Column F
    cell_c = ws.cell(row_idx, 3)  # Column C (base hours formula)
    
    if cell_f.value and isinstance(cell_f.value, str) and cell_f.value.startswith('=IF'):
        print(f'Row {row_idx}: {category_name}')
        print(f'  Column C (Base): {cell_c.value}')
        print(f'  Column F (Tier Adjustment):')
        # Print formula in multiple lines for readability
        formula = str(cell_f.value)
        # Find the named range being used
        if 'Project_Initiation' in formula:
            named_range = 'Project_Initiation_and_Planning'
        elif 'Requirement_Gathering' in formula:
            named_range = 'Requirement_Gathering'
        elif 'Build_and_Configure' in formula:
            named_range = 'Build_and_Configure'
        elif 'AppFeatures' in formula:
            named_range = 'AppFeatures'
        elif 'AppCustomization' in formula:
            named_range = 'AppCustomization'
        elif 'Calculations' in formula:
            named_range = 'Calculations'
        elif 'Security' in formula:
            named_range = 'Security'
        elif 'HistData' in formula:
            named_range = 'HistData'
        elif 'Integrations' in formula:
            named_range = 'Integrations'
        elif 'Reporting' in formula:
            named_range = 'Reporting'
        elif 'Automations' in formula:
            named_range = 'Automations'
        elif 'Testing_Training' in formula:
            named_range = 'Testing_Training'
        elif 'Transition' in formula:
            named_range = 'Transition'
        elif 'Documentations' in formula:
            named_range = 'Documentations'
        elif 'Change_Management' in formula:
            named_range = 'Change_Management'
        elif 'Design' in formula:
            named_range = 'Design'
        else:
            named_range = 'UNKNOWN'
        
        print(f'  Formula structure: IF(W<=100, {named_range}, IF(W<=120, {named_range}+X, IF(W<=160, {named_range}+Y, {named_range}+Z)))')
        print(f'  Full formula: {formula[:150]}...')
        print()
