"""
Extract actual calculated values from Excel for complex formula tasks
"""
import openpyxl
from pathlib import Path

excel_path = Path("C:\\Users\\NavashanmugamAsokan\\Desktop\\Projects\\Engagement Scoping Tool\\scoping_tool\\Engagement Scoping Tool - FCC.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['Effort Estimation']

print("=" * 120)
print("COMPLEX FORMULA TASKS - CALCULATED VALUES")
print("=" * 120)
print(f"{'Row':>3} {'Task':45s} {'C (Base)':>8} {'E (Details)':>10} {'F (Final)':>12} {'Formula Result Notes':>30}")
print("-" * 120)

# Check specific tasks with formulas
tasks_to_check = {
    71: "Data Forms",
    72: "Dashboards",
    75: "Business Rules",
    79: "Management Reports",
}

for row, task_name in tasks_to_check.items():
    col_b = ws[f'B{row}'].value
    col_c = ws[f'C{row}'].value
    col_e = ws[f'E{row}'].value
    col_f = ws[f'F{row}'].value
    
    print(f"{row:3d} {str(col_b)[:44] if col_b else task_name:45s} {str(col_c):>8} {str(col_e):>10} {str(col_f):>12}")

print("\n" + "=" * 120)
print("CHECKING: Does Column E contain user input or calculated weightage?")
print("=" * 120)

# Check what's in the Scope Definition sheet for Data Forms
wb_scope = openpyxl.load_workbook(excel_path, data_only=True)
ws_scope = wb_scope['Scope Definition']

print("\nScope Definition Sheet - Data Forms (Row 40):")
print(f"  B40: {ws_scope['B40'].value}")  # Name
print(f"  C40: {ws_scope['C40'].value}")  # In Scope?
print(f"  D40: {ws_scope['D40'].value}")  # Details input
print(f"  E40: {ws_scope['E40'].value}")  # Calculated weightage
