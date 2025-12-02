"""
Example: Load scope from Excel and run complete workflow
"""

from pathlib import Path
from openpyxl import load_workbook
import sys

sys.path.insert(0, str(Path(__file__).parent))

from backend.scoping_engine import ScopingEngine
from backend.config import EXCEL_FILE, SHEET_SCOPE_DEFINITION


def load_scope_from_excel():
    """Load actual scope data from Excel for testing"""
    print("Loading scope data from Excel...")
    
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_SCOPE_DEFINITION]
    
    scope_inputs = []
    
    for row in range(6, 103):
        metric_name = ws.cell(row=row, column=2).value
        if metric_name and isinstance(metric_name, str) and metric_name.strip():
            # Skip headings (no leading spaces)
            if not metric_name.startswith(' '):
                continue
            
            in_scope_val = ws.cell(row=row, column=3).value
            details_val = ws.cell(row=row, column=4).value  # Column D (index 4) is Details
            
            scope_inputs.append({
                'name': metric_name.strip(),
                'in_scope': str(in_scope_val).upper() if in_scope_val else 'NO',
                'details': int(details_val) if details_val and isinstance(details_val, (int, float)) else 0
            })
    
    print(f"✓ Loaded {len(scope_inputs)} metrics from Excel\n")
    
    return {
        'scope_inputs': scope_inputs,
        'selected_roles': ['PM USA', 'PM India', 'Architect USA']
    }


def main():
    """Run complete scoping workflow with Excel data"""
    
    # Load inputs from Excel
    user_input = load_scope_from_excel()
    
    # Initialize scoping engine
    engine = ScopingEngine()
    
    # Run complete workflow
    report = engine.run_complete_workflow(user_input, 'test_scoping_report.json')
    
    print("\n✓ Workflow completed successfully!")
    print(f"✓ Report generated with {len(report['scope_definition']['metrics'])} metrics analyzed")


if __name__ == "__main__":
    main()
