"""
Extract all category totals from Column G (which should be the summary per category)
"""
import openpyxl
from pathlib import Path

excel_path = Path("C:\\Users\\NavashanmugamAsokan\\Desktop\\Projects\\Engagement Scoping Tool\\scoping_tool\\Engagement Scoping Tool - FCC.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['Effort Estimation']

print("=" * 100)
print("CATEGORY TOTALS FROM COLUMN G (G = F + Tasks)")
print("=" * 100)
print(f"{'Row':>3} {'Category':45s} {'G (Total)':>15}")
print("-" * 100)

# Scan all rows to find category headers (those with formulas in F and G)
total = 0
for row in range(5, 130):
    col_b = ws[f'B{row}'].value
    col_g = ws[f'G{row}'].value
    
    # Check if this row has a value in column G that looks like a number/formula
    if col_b and col_g and isinstance(col_g, (int, float)) and col_g > 0:
        print(f"{row:3d} {str(col_b)[:44]:45s} {col_g:>15.1f}")
        total += col_g

print("-" * 100)
print(f"{'TOTAL':>51} {total:>15.1f}")
print("=" * 100)
print(f"\nExpected from image: 1892.5 hours")
print(f"Calculated from Excel Column G sum: {total} hours")
print(f"Match: {abs(total - 1892.5) < 1}")
