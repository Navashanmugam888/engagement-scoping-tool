"""
Get exact formulas for Data Forms and Dashboards
"""
import openpyxl
from pathlib import Path

excel_path = Path("C:\\Users\\NavashanmugamAsokan\\Desktop\\Projects\\Engagement Scoping Tool\\scoping_tool\\Engagement Scoping Tool - FCC.xlsx")
wb = openpyxl.load_workbook(excel_path)
ws = wb['Effort Estimation']

print("=" * 120)
print("EXACT FORMULAS")
print("=" * 120)

for row in [71, 72]:
    cell_b = ws[f'B{row}'].value
    cell_f = ws[f'F{row}'].value
    
    print(f"\nRow {row}: {cell_b}")
    print(f"F{row} formula: {cell_f}")
