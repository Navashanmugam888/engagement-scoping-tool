import openpyxl

excel_file = 'C:/Users/NavashanmugamAsokan/Desktop/Projects/Engagement Scoping Tool/documents/Engagement Scoping Tool - FCC.xlsx'

wb = openpyxl.load_workbook(excel_file, data_only=False)
ws_scope = wb['Scope Definition']
ws_effort = wb['Effort Estimation']

print('=== SCOPE DEFINITION - LOOKING FOR TIER/WEIGHTAGE ===')
print()

# Look for summary rows
for row_idx in range(130, 155):
    row = ws_scope[row_idx]
    values = []
    for col_idx in range(1, 6):
        cell = row[col_idx-1]
        if cell.value:
            val = str(cell.value)
            if len(val) > 50:
                val = val[:50] + '...'
            values.append(f'{chr(64+col_idx)}{row_idx}: {val}')
    if values:
        print(f'Row {row_idx}:')
        for v in values:
            print(f'  {v}')

print()
print('=== EFFORT ESTIMATION - Sample Tier Adjustment Formula (Row 5) ===')
print()
cell_f5 = ws_effort.cell(5, 6)  # Final Estimate column
print(f'Row 5, Col F (Final Estimate):')
print(f'  Formula: {cell_f5.value}')
print()

print('=== EFFORT ESTIMATION - Row 18 Example ===')
print()
for col in range(1, 9):
    cell = ws_effort.cell(18, col)
    print(f'  Col{col}: {cell.value}')

print()
print('=== Checking if there is Tier lookup table ===')
ws_app_tiers = wb['App Tiers Definition']
print('App Tiers Definition sheet rows:')
for row_idx in range(1, 20):
    row = ws_app_tiers[row_idx]
    values = []
    for col_idx in range(1, 5):
        cell = row[col_idx-1]
        if cell.value:
            values.append(f'{chr(64+col_idx)}: {str(cell.value)[:30]}')
    if values:
        print(f'Row {row_idx}: {", ".join(values)}')
