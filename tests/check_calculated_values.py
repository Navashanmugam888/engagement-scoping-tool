"""
Get actual calculated values for rows 71, 72
"""
import openpyxl
from pathlib import Path

excel_path = Path("C:\\Users\\NavashanmugamAsokan\\Desktop\\Projects\\Engagement Scoping Tool\\scoping_tool\\Engagement Scoping Tool - FCC.xlsx")

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['Effort Estimation']

print("=" * 100)
print("CALCULATED VALUES (data_only=True)")
print("=" * 100)

for row in [70, 71, 72, 73]:
    col_b = ws[f'B{row}'].value
    col_f = ws[f'F{row}'].value
    col_g = ws[f'G{row}'].value
    
    print(f"Row {row}: {str(col_b):40s} F={col_f} G={col_g}")
