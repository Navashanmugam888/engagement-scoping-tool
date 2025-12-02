"""
Examine Historical Data section and its tasks
"""
import openpyxl
from pathlib import Path

excel_path = Path("C:\\Users\\NavashanmugamAsokan\\Desktop\\Projects\\Engagement Scoping Tool\\scoping_tool\\Engagement Scoping Tool - FCC.xlsx")
wb = openpyxl.load_workbook(excel_path)
ws = wb['Effort Estimation']

print("=" * 130)
print("HISTORICAL DATA SECTION (Rows 84-89)")
print("=" * 130)
print(f"{'Row':>3} {'B: Task':45s} {'C: Base':>7} {'E: Details':>12} {'F: Final Est':>20} {'Formula in F':>50}")
print("-" * 130)

for row in range(84, 89):
    col_b = ws[f'B{row}'].value or ""
    col_c = ws[f'C{row}'].value or ""
    col_e = ws[f'E{row}'].value or ""
    col_f_val = ws[f'F{row}'].value or ""
    
    # Get formula if cell contains one
    cell_f = ws[f'F{row}']
    col_f_formula = cell_f.value if cell_f.data_type == 'f' else ""
    
    print(f"{row:3d} {str(col_b)[:44]:45s} {str(col_c):>7} {str(col_e):>12} {str(col_f_val):>20} {str(col_f_formula)[:48]:>50}")

print("\n" + "=" * 130)
print("CHECKING FOR NAMED RANGES IN FORMULAS")
print("=" * 130)

# Check the actual formula content
for row in [84, 85, 86, 87, 88]:
    cell = ws[f'F{row}']
    if cell.value and isinstance(cell.value, str):
        print(f"F{row}: {cell.value}")
