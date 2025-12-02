"""
Extract exact effort values from Excel for each category to compare
"""
import openpyxl
from pathlib import Path
import re

# Load Excel workbook
excel_path = Path("C:\\Users\\NavashanmugamAsokan\\Desktop\\Projects\\Engagement Scoping Tool\\scoping_tool\\Engagement Scoping Tool - FCC.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['Effort Estimation']

print("=" * 100)
print("EXCEL EFFORT ESTIMATION VALUES (Data Only)")
print("=" * 100)

for row in range(5, 50):
    col_b = ws[f'B{row}'].value  # Task/Category
    col_c = ws[f'C{row}'].value  # Base Hours
    col_f = ws[f'F{row}'].value  # Final Estimate
    
    if col_b:
        print(f"Row {row:2d}: {str(col_b):50s} Base: {str(col_c):7s} -> Final: {str(col_f):7s}")
    if row == 35:
        print("-" * 100)
