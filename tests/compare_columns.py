"""
Compare Column F vs Column G for categories
"""
import openpyxl
from pathlib import Path

excel_path = Path("C:\\Users\\NavashanmugamAsokan\\Desktop\\Projects\\Engagement Scoping Tool\\scoping_tool\\Engagement Scoping Tool - FCC.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['Effort Estimation']

print("=" * 100)
print("CATEGORY FINAL ESTIMATES: Column F vs Column G")
print("=" * 100)
print(f"{'Row':>3} {'Category':40s} {'F (Base+Tier)':>15} {'G (SUM)':>15} {'Difference':>15}")
print("-" * 100)

# Check category header rows
category_rows = [5, 14, 18, 28, 35, 52, 66, 82, 90, 101, 109, 117, 125]

for row in category_rows:
    col_b = ws[f'B{row}'].value
    col_f = ws[f'F{row}'].value
    col_g = ws[f'G{row}'].value
    
    if col_b:
        diff = (col_g or 0) - (col_f or 0) if col_f and col_g else None
        print(f"{row:3d} {str(col_b)[:39]:40s} {str(col_f if col_f else 'N/A'):>15} {str(col_g if col_g else 'N/A'):>15} {str(diff if diff else 'N/A'):>15}")
