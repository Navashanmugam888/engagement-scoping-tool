import openpyxl

excel_file = 'C:/Users/NavashanmugamAsokan/Desktop/Projects/Engagement Scoping Tool/documents/Engagement Scoping Tool - FCC.xlsx'

wb = openpyxl.load_workbook(excel_file, data_only=False)
ws = wb['Effort Estimation']

print('=== AUTOMATIONS SECTION ===')
print()
for row_idx in range(106, 111):
    print(f'Row {row_idx}:')
    for col in range(1, 9):
        cell = ws.cell(row_idx, col)
        col_letter = chr(64+col)
        if cell.value:
            val = str(cell.value)[:80]
            print(f'  {col_letter}: {val}')
    print()

print('=== TESTING/TRAINING SECTION ===')
print()
for row_idx in range(112, 118):
    print(f'Row {row_idx}:')
    for col in range(1, 9):
        cell = ws.cell(row_idx, col)
        col_letter = chr(64+col)
        if cell.value:
            val = str(cell.value)[:80]
            print(f'  {col_letter}: {val}')
    print()

print('=== TRANSITION SECTION ===')
print()
for row_idx in range(119, 122):
    print(f'Row {row_idx}:')
    for col in range(1, 9):
        cell = ws.cell(row_idx, col)
        col_letter = chr(64+col)
        if cell.value:
            val = str(cell.value)[:80]
            print(f'  {col_letter}: {val}')
    print()
