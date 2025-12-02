import openpyxl

excel_file = 'C:/Users/NavashanmugamAsokan/Desktop/Projects/Engagement Scoping Tool/documents/Engagement Scoping Tool - FCC.xlsx'

wb = openpyxl.load_workbook(excel_file, data_only=False)
ws = wb['Effort Estimation']

print('=== TASKS WITH DETAILS FORMULAS vs TASKS WITHOUT ===')
print()

print('TASKS WITHOUT F-COLUMN FORMULAS (no detail multiplication):')
no_formula = []
for row_idx in range(36, 52):
    cell_b = ws.cell(row_idx, 2)
    cell_f = ws.cell(row_idx, 6)
    if cell_b.value and (not cell_f.value or cell_f.value == ''):
        no_formula.append((row_idx, str(cell_b.value)))

for row, task in no_formula:
    print(f'  Row {row}: {task}')

print()
print('TASKS WITH F-COLUMN FORMULAS (detail multiplication):')
with_formula = []
for row_idx in range(36, 52):
    cell_b = ws.cell(row_idx, 2)
    cell_f = ws.cell(row_idx, 6)
    if cell_b.value and cell_f.value and isinstance(cell_f.value, str) and '=' in cell_f.value:
        with_formula.append((row_idx, str(cell_b.value), str(cell_f.value)[:60]))

for row, task, formula in with_formula:
    print(f'  Row {row}: {task}')
    print(f'    Formula: {formula}')

print()
print('KEY FINDING:')
print('Tasks without formulas contribute 0 to Final Estimate (Column F)')
print('Tasks with formulas multiply base hours by details from Scope Definition')
print()
print('So for "Build and Configure FCC" category:')
print('- Final Estimate = SUM(F35 + subtask_F_values)')
print('- Where F35 is the tier-adjusted category base')
print('- And subtask_F_values only exist for tasks with detail requirements')
