"""
Get the exact formula for Data Forms
"""
import openpyxl
from pathlib import Path

excel_path = Path("C:\\Users\\NavashanmugamAsokan\\Desktop\\Projects\\Engagement Scoping Tool\\scoping_tool\\Engagement Scoping Tool - FCC.xlsx")
wb = openpyxl.load_workbook(excel_path)
ws = wb['Effort Estimation']

print("Data Forms (Row 71) - Formula Analysis:")
print("=" * 100)

# Get the formula
cell_f71 = ws['F71']
print(f"\nFormula in F71: {cell_f71.value}")
print(f"Data type: {cell_f71.data_type}")

# Check if it's an array formula
if hasattr(cell_f71, '_value'):
    print(f"Direct value: {cell_f71._value}")

# Get values from nearby cells
print(f"\nE71 (should be 5): {ws['E71'].value}")
print(f"C71 (base hours): {ws['C71'].value}")
print(f"D71 (in scope?): {ws['D71'].value}")

# Also check the actual cell range
print(f"\nCell F71 details:")
print(f"  Value: {ws['F71'].value}")
print(f"  Type: {type(ws['F71'].value)}")

# Check with data_only
wb_data = openpyxl.load_workbook(excel_path, data_only=True)
ws_data = wb_data['Effort Estimation']
print(f"\nWith data_only=True:")
print(f"  F71 calculated value: {ws_data['F71'].value}")
print(f"  E71 value: {ws_data['E71'].value}")
