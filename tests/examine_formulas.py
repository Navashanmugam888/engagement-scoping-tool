#!/usr/bin/env python
"""Examine Excel formula structure for categories"""

import openpyxl

wb = openpyxl.load_workbook('Engagement Scoping Tool - FCC.xlsx')
ws = wb['Effort Estimation']

# Critical rows - category headers
rows_to_check = [
    (5, 'Project Initiation and Planning'),
    (35, 'Build and Configure FCC'),
    (74, 'Calculations'),
    (81, 'Secured Dimensions'),
    (84, 'Historical Data'),
    (90, 'Integrations'),
    (98, 'Management Reports'),
]

print('EXAMINING EXCEL FORMULAS FOR CATEGORIES')
print('=' * 120)

for row, cat_name in rows_to_check:
    print()
    print('Row {}: {}'.format(row, cat_name))
    print('-' * 120)
    
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        cell = ws[col_letter + str(row)]
        print('  {}{}: Value={:20s} Formula={}'.format(col_letter, row, str(cell.value)[:20], cell.value))

print()
print('=' * 120)
print('Now checking subtask structure for Build and Configure FCC (rows 36-51):')
print('=' * 120)

for row in range(36, 52):
    col_b = ws[f'B{row}'].value
    col_c = ws[f'C{row}'].value
    col_e = ws[f'E{row}'].value
    col_f = ws[f'F{row}'].value
    
    if col_b:
        print('Row {:3d}: {} | C={:>6s} | E={:>6s} | F={:>6s}'.format(
            row, str(col_b)[:50], str(col_c)[:6], str(col_e)[:6], str(col_f)[:6]))
