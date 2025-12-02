"""
Extract all category header rows and their formulas
"""
import openpyxl
from pathlib import Path

# Load Excel workbook
excel_path = Path("C:\\Users\\NavashanmugamAsokan\\Desktop\\Projects\\Engagement Scoping Tool\\scoping_tool\\Engagement Scoping Tool - FCC.xlsx")
wb = openpyxl.load_workbook(excel_path)
ws = wb['Effort Estimation']

print("=" * 120)
print("CATEGORY HEADERS - FORMULAS IN COLUMN F (Final Estimate)")
print("=" * 120)

# Check specific rows that are likely category headers
category_rows = {
    5: "Project Initiation and Planning",
    14: "Creating and Managing EPM Cloud Infrastructure",
    18: "Requirement Gathering",
    28: "Design",
    35: "Build and Configure FCC",
}

for row_num, category_name in category_rows.items():
    cell_b = ws[f'B{row_num}'].value
    cell_c = ws[f'C{row_num}'].value
    cell_f = ws[f'F{row_num}'].value
    
    # Try to get the formula
    cell_formula = ws[f'F{row_num}'].value if isinstance(ws[f'F{row_num}'].value, str) else ""
    
    print(f"\nRow {row_num}:")
    print(f"  Category: {cell_b}")
    print(f"  Base (C{row_num}): {cell_c}")
    print(f"  Final Est (F{row_num}): {cell_f}")
    print(f"  Formula (F{row_num}): {ws[f'F{row_num}'].value if ws[f'F{row_num}'].data_type == 'f' else 'NO FORMULA'}")

# Now let's read the raw formulas
print("\n" + "=" * 120)
print("RAW FORMULAS")
print("=" * 120)

wb_formula = openpyxl.load_workbook(excel_path)
ws_formula = wb_formula['Effort Estimation']

for row_num in [5, 14, 18, 28, 35]:
    cell = ws_formula[f'F{row_num}']
    print(f"F{row_num}: {cell.value}")