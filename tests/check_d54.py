"""
Check what's in Scope Definition D54
"""
import openpyxl
from pathlib import Path

excel_path = Path("C:\\Users\\NavashanmugamAsokan\\Desktop\\Projects\\Engagement Scoping Tool\\scoping_tool\\Engagement Scoping Tool - FCC.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['Scope Definition']

print("Checking rows 50-60 in Scope Definition:")
print("=" * 100)
print(f"{'Row':>3} {'B':40s} {'C':>10} {'D':>12} {'E':>12}")
print("-" * 100)

for row in range(50, 61):
    col_b = ws[f'B{row}'].value
    col_c = ws[f'C{row}'].value
    col_d = ws[f'D{row}'].value
    col_e = ws[f'E{row}'].value
    if col_b or col_d:
        print(f"{row:3d} {str(col_b)[:39] if col_b else '':40s} {str(col_c) if col_c else '':>10} {str(col_d) if col_d else '':>12} {str(col_e) if col_e else '':>12}")
