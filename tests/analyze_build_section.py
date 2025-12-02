"""
Check the full structure of the Build and Configure section to understand
whether task estimates should be included in category final estimate
"""
import openpyxl
from pathlib import Path

excel_path = Path("C:\\Users\\NavashanmugamAsokan\\Desktop\\Projects\\Engagement Scoping Tool\\scoping_tool\\Engagement Scoping Tool - FCC.xlsx")
wb = openpyxl.load_workbook(excel_path)
ws = wb['Effort Estimation']

print("=" * 120)
print("BUILD AND CONFIGURE FCC SECTION (Rows 35-51)")
print("=" * 120)
print(f"{'Row':>3} {'Column B':50s} {'C (Base)':>7} {'E (Details)':>7} {'F (Final)':>12} {'G (Category)':>12}")
print("-" * 120)

for row in range(35, 52):
    col_b = ws[f'B{row}'].value or ""
    col_c = ws[f'C{row}'].value or ""
    col_e = ws[f'E{row}'].value or ""
    col_f = ws[f'F{row}'].value or ""
    col_g = ws[f'G{row}'].value or ""
    
    print(f"{row:3d} {str(col_b)[:49]:50s} {str(col_c):>7} {str(col_e):>7} {str(col_f):>12} {str(col_g):>12}")
