import openpyxl

excel_file = 'C:/Users/NavashanmugamAsokan/Desktop/Projects/Engagement Scoping Tool/documents/Engagement Scoping Tool - FCC.xlsx'

wb = openpyxl.load_workbook(excel_file, data_only=False)
ws = wb['Effort Estimation']

print('=== EFFORT ESTIMATION SHEET - DETAILED STRUCTURE ===')
print()
print('Row structure analysis:')
print()

# Analyze category rows
print('CATEGORY ROWS (with tier-adjusted Final Estimate):')
print()

category_rows = [5, 14, 18, 30, 39, 51, 60, 76, 82, 86, 91, 95, 100, 107, 119, 123, 128]

for row_idx in category_rows:
    row = ws[row_idx]
    print(f'Row {row_idx}:')
    for col in range(1, 9):
        cell = row[col - 1]
        val = cell.value
        if val:
            if isinstance(val, str) and val.startswith('='):
                print(f'  {chr(64+col)}: [FORMULA] {val[:80]}')
            else:
                print(f'  {chr(64+col)}: {val}')
    print()

print('=' * 80)
print('SUMMARY ROW (Row 131-132):')
print()

for row_idx in [131, 132]:
    row = ws[row_idx]
    print(f'Row {row_idx}:')
    for col in range(6, 9):  # Columns F, G, H
        cell = row[col - 1]
        val = cell.value
        col_letter = chr(64+col)
        if val:
            if isinstance(val, str) and val.startswith('='):
                print(f'  {col_letter}: [FORMULA] {val}')
            else:
                print(f'  {col_letter}: {val}')
    print()

print('=' * 80)
print('UNDERSTANDING:')
print()
print('F131: "Hours" (header)')
print('F132: =SUM(G5:G130) - This sums the Final Estimate (Column G) NOT Column F!')
print()
print('Wait - let me check if columns are shifted...')
print()

# Check what columns actually contain
print('Row 5 (Project Initiation and Planning):')
for col in range(1, 9):
    cell = ws.cell(5, col)
    col_letter = chr(64+col)
    if cell.value:
        val = str(cell.value)[:60]
        print(f'  {col_letter}: {val}')

print()
print('Row 6 (Kickoff Meetings - subtask):')
for col in range(1, 9):
    cell = ws.cell(6, col)
    col_letter = chr(64+col)
    if cell.value:
        val = str(cell.value)[:60]
        print(f'  {col_letter}: {val}')
