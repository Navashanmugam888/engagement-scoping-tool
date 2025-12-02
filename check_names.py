from openpyxl import load_workbook

wb = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=False)

# Check defined names
print('Defined names in workbook:')
if hasattr(wb, 'defined_names'):
    for name_obj in wb.defined_names.keys():
        print(f'Name: {name_obj}')
        defined = wb.defined_names[name_obj]
        print(f'  Value: {defined}')

ws = wb['Effort Estimation']

# Check E84 specifically - this might have the formula for EngagementWeightage
print('\n\nRow 84 cells:')
for col in ['D', 'E', 'F']:
    cell = ws[f'{col}84']
    print(f'{col}84: value={cell.value}, formula={cell.value if not hasattr(cell, "value") else cell.value}')
