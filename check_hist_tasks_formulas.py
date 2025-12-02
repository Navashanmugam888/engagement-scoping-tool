from openpyxl import load_workbook

wb = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=False)
ws = wb['Effort Estimation']

print('Looking at Historical Data tasks (rows 85-88):')
for row in range(85, 89):
    print(f'\nRow {row}:')
    print(f'  B: {ws[f"B{row}"].value}')
    print(f'  C: {ws[f"C{row}"].value}')
    print(f'  D: {ws[f"D{row}"].value}')
    print(f'  E: {ws[f"E{row}"].value}')
    print(f'  F: {ws[f"F{row}"].value}')

# Check with data values
wb2 = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=True)
ws2 = wb2['Effort Estimation']

print('\n\nWith calculated values:')
for row in range(85, 89):
    print(f'\nRow {row}:')
    print(f'  B: {ws2[f"B{row}"].value}')
    print(f'  C: {ws2[f"C{row}"].value}')
    print(f'  D: {ws2[f"D{row}"].value}')
    print(f'  E: {ws2[f"E{row}"].value}')
    print(f'  F: {ws2[f"F{row}"].value}')
