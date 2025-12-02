from openpyxl import load_workbook

# Load Excel with data_only=True to see calculated values
wb = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=True)
ws = wb['Effort Estimation']

# Check row 84 (Historical Data category)
print('Row 84 - Historical Data category:')
print(f'  B84: {ws["B84"].value}')
print(f'  C84 (Base): {ws["C84"].value}')
print(f'  F84 (Final with tier): {ws["F84"].value}')
print(f'  G84 (Sum of tasks): {ws["G84"].value}')

# Also check rows 85-88 (tasks)
print('\nTasks in Historical Data category (rows 85-88):')
for row in range(85, 89):
    print(f'  Row {row}:')
    print(f'    B: {ws[f"B{row}"].value}')
    print(f'    C: {ws[f"C{row}"].value}')
    print(f'    F: {ws[f"F{row}"].value}')
