from openpyxl import load_workbook

wb = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=False)
ws = wb['Effort Estimation']

# Find Historical Data category row (should be around row 71-73)
print('Looking for Historical Data category row...')
for row in range(70, 90):
    b_val = ws[f'B{row}'].value
    if b_val and str(b_val).strip() == 'Historical Data':
        print(f'\nRow {row}: {b_val}')
        print(f'  C{row}: {ws[f"C{row}"].value}')
        print(f'  D{row}: {ws[f"D{row}"].value}')
        print(f'  E{row}: {ws[f"E{row}"].value}')
        print(f'  F{row}: {ws[f"F{row}"].value}')
        print(f'  G{row}: {ws[f"G{row}"].value}')

# Also check with data_only to see calculated values
wb2 = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=True)
ws2 = wb2['Effort Estimation']

print('\n\nWith calculated values:')
for row in range(70, 90):
    b_val = ws2[f'B{row}'].value
    if b_val and str(b_val).strip() == 'Historical Data':
        print(f'\nRow {row}: {b_val}')
        print(f'  C{row}: {ws2[f"C{row}"].value}')
        print(f'  D{row}: {ws2[f"D{row}"].value}')
        print(f'  E{row}: {ws2[f"E{row}"].value}')
        print(f'  F{row}: {ws2[f"F{row}"].value}')
        print(f'  G{row}: {ws2[f"G{row}"].value}')
