from openpyxl import load_workbook

wb = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=False)
ws = wb['Scope Definition']

print('Scope Definition sheet, row 103:')
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    cell = ws[f'{col}103']
    print(f'{col}103: {cell.value}')

# Also check with data_only
print('\n\nWith calculated values:')
wb2 = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=True)
ws2 = wb2['Scope Definition']

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    cell = ws2[f'{col}103']
    print(f'{col}103: {cell.value}')
