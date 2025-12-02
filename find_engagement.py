from openpyxl import load_workbook

wb = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=False)

# Check defined names
print('Defined names in workbook:')
if hasattr(wb, 'defined_names'):
    for name in wb.defined_names.definedName:
        print(f'{name.name}: {name.value}')

ws = wb['Effort Estimation']

# Search for EngagementWeightage in formulas
print('\n\nSearching for "EngagementWeightage" in formulas...')
for row in range(1, 100):
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws[f'{col}{row}']
        if cell.value and isinstance(cell.value, str) and 'EngagementWeightage' in cell.value:
            print(f'{col}{row}: {cell.value}')

# Check row 84 column E specifically
print('\n\nChecking row 84:')
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    print(f'{col}84: {ws[f"{col}84"].value}')
