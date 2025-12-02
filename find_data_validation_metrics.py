from openpyxl import load_workbook

wb = load_workbook('Engagement Scoping Tool - FCC.xlsx', data_only=True)
ws = wb['Scope Definition']

# Find the data validation metrics
print('Looking for data validation metrics in Scope Definition:')
for row in range(6, 103):
    b_val = ws[f'B{row}'].value
    if b_val and isinstance(b_val, str):
        if 'Data Validation' in b_val or 'Historical Journal' in b_val:
            print(f'\nRow {row}: {b_val}')
            print(f'  C (In Scope?): {ws[f"C{row}"].value}')
            print(f'  D (Details): {ws[f"D{row}"].value}')
            print(f'  E (Weightage): {ws[f"E{row}"].value}')
