import openpyxl

wb = openpyxl.load_workbook('Engagement Scoping Tool - FCC.xlsx')
ws = wb['Effort Estimation']

print("=" * 80)
print("HISTORICAL DATA SECTION - Excel Formulas")
print("=" * 80)

print("\n1. Parent Task:")
print(f"Row 83: {ws['A83'].value}")
print(f"  Column C (Base Hours): {ws['C83'].value}")
print(f"  Column E (Formula): {ws['E83'].value}")

print("\n2. Child Tasks:")
print(f"\nRow 86: {ws['A86'].value}")
print(f"  Column C (Base Hours): {ws['C86'].value}")
print(f"  Column E (Formula): {ws['E86'].value}")

print(f"\nRow 87: {ws['A87'].value}")
print(f"  Column C (Base Hours): {ws['C87'].value}")
print(f"  Column E (Formula): {ws['E87'].value}")

print(f"\nRow 88: {ws['A88'].value}")
print(f"  Column C (Base Hours): {ws['C88'].value}")
print(f"  Column E (Formula): {ws['E88'].value}")

print("\n" + "=" * 80)
print("EXPLANATION:")
print("=" * 80)
print("\nRow 86 & 87 have base hours = 96")
print("  This is 12 hours/day × 8 days = 96 hours")
print("  Formula multiplies 96 by the details from Historical Data Validation")
print("\nRow 88 has base hours = 8")
print("  Formula multiplies 8 by the details from Historical Data Validation")
